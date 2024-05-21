#Import the module openpyxl
import openpyxl

#Import the Excel sales data
import csv

#Read the data from the spreadsheet
def read_data():
    data = []

    with open("sales.csv", "r") as sales_csv:
        spreadsheet = csv.DictReader(sales_csv)
        for row in spreadsheet:
            data.append(row)

    return data

def run():
    data = read_data()

    sales = []
    dictionary = {}

    for row in data:
        sale = int(row["sales"])
        month = row["month"]
        sales.append(sale)

        dictionary[sale] = month

#Collect all the sales from each month into a single list
    list_of_sales = list(sales)
    print("List of sales: {}".format(list_of_sales))
    print()

#Show the monthly changes as a percentage and fill NaN with a value
    import pandas as pd
    df = pd.read_csv(r"sales.csv")
    df["percentage-change for sales"] = df["sales"].pct_change() * 100
    df = df.fillna(0)
    print(df)

    output_df = pd.DataFrame()

#Create an excelwriter object
    with pd.ExcelWriter(r"Book_Sales.xlsx") as writer:
        df.to_excel(writer, sheet_name="Results Sheet", index=True)
        output_df.to_excel(writer, sheet_name="Results Sheet", index=False)

#Load the existing workbook
    from openpyxl import load_workbook
    workbook_name = "Book_Sales.xlsx"
    wb = load_workbook(workbook_name)
    page = wb.active

#Set the column dimensions as best fit
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook("Book_Sales.xlsx")
    ws = wb["Results Sheet"]
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True

#Save the changes
    wb.save("Book_Sales.xlsx")
    wb.close()

    from openpyxl import load_workbook
    wb2 = load_workbook("Book_Sales.xlsx")
    wb2.create_sheet("Yearly Results")
    wb2.save("Book_Sales.xlsx")
    wb.close()

#Output the total sales across all months
    total = sum(sales)
    print("\n""The total value of sales is £{}".format(total))

#Find the maximum and minimum values of sales
    max_value = max(sales)
    min_value = min(sales)

#Find the months with the lowest and highest number of sales
    max_month = dictionary.get(max_value)
    min_month = dictionary.get(min_value)

#Find the average value of sales
    number_of_sales = len(sales)
    average_sales_value = total / number_of_sales

    import math

    number = average_sales_value
    rounded_number = math.ceil(number)

#Print these values
    print("The maximum value is £{}".format(max_value))
    print("The minimum value is £{}".format(min_value))
    print("The month with the highest number of sales is {}".format(max_month))
    print("The month with the lowest number of sales is {}".format(min_month))
    print("The average value is £{}".format(rounded_number))

    import pandas as pd

#Create a dataframe for the new data
    df = pd.DataFrame({"Sales Total": [total], "Sales Average": [rounded_number], "Sales Maximum": [max_value], "Sales Minimum": [min_value], "Month with Highest Sales": [max_month], "Month with Lowest Sales": [min_month]})

#Use ExcelWriter with append mode and replace option; index=False removes surplus numbering from the first column
    with pd.ExcelWriter(r"Book_Sales.xlsx", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Yearly Results", index=False)

#Save the workbook to keep the changes
    wb.save("Book_Sales.xlsx")
    wb.close()

run()